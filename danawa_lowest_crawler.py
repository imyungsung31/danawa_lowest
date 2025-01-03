from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from datetime import datetime
import pytz
import os
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.utils.dataframe import dataframe_to_rows
from selenium.webdriver.chrome.options import Options
import streamlit as st


# pcode = "70531547"
pcode = "69059459"

def crawl_and_update_excel():
    # Chrome 드라이버 설정
    chrome_options = Options()
    chrome_options.add_argument("--headless")  # headless 모드 활성화
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    driver = webdriver.Chrome(options=chrome_options)

    # 웹페이지 로드
    url = "https://prod.danawa.com/info/?pcode=" + pcode
    driver.get(url)

    # 현재 시간의 가격 데이터를 저장할 딕셔너리
    kst = pytz.timezone('Asia/Seoul')
    current_time = datetime.now(kst).strftime('%Y-%m-%d %H:%M')
    current_prices = {'날짜 및 시간': current_time}

    try:
        # 명시적 대기 설정
        wait = WebDriverWait(driver, 10)

        # lowest list tr개수 가져오기 
        browser = driver
        xpath = "/html/body/div[2]/div[5]/div[2]/div[2]/div[2]/div[1]/div[2]/div[3]/table/tbody[1]/tr"
        tr_elements = browser.find_elements(By.XPATH, xpath)
        rows = len(tr_elements)
        print(f"총 {rows}개의 항목이 있습니다.")
        
        for i in range(1, rows + 1):
            try:
                # 이미지 요소 찾기 및 src 속성 가져오기
                img_xpath = f"/html/body/div[2]/div[5]/div[2]/div[2]/div[2]/div[1]/div[2]/div[3]/table/tbody[1]/tr[{i}]/td[1]/div/a/img"
                img_element = wait.until(EC.presence_of_element_located((By.XPATH, img_xpath)))
                img_alt = img_element.get_attribute('alt')
            except:
                # 이미지 요소가 없을 경우 텍스트 추출
                img_xpath = f"/html/body/div[2]/div[5]/div[2]/div[2]/div[2]/div[1]/div[2]/div[3]/table/tbody[1]/tr[{i}]/td[1]/div/a"
                img_alt = wait.until(EC.presence_of_element_located((By.XPATH, img_xpath))).text
            
            # 가격 요소 찾기 및 텍스트 가져오기
            price_xpath = f"/html/body/div[2]/div[5]/div[2]/div[2]/div[2]/div[1]/div[2]/div[3]/table/tbody[1]/tr[{i}]/td[2]/a/span/em"
            price_element = wait.until(EC.presence_of_element_located((By.XPATH, price_xpath)))
            price_text = price_element.text.replace(',', '')  # 쉼표 제거
            price_text = float(price_text)  # 숫자로 변환
            
            # 현재 시간의 가격 데이터에 추가
            current_prices[img_alt] = price_text
            print(f"MALL {i}: {img_alt}")
            print(f"가격 {i}: {price_text}")

    except Exception as e:
        print("에러 발생:", str(e))

    finally:
        # 브라우저 종료
        driver.quit()

        file_path = 'danawa_lowest_prices_' + pcode + '.xlsx'
        
        # 현재 가격 데이터를 DataFrame으로 변환
        new_df = pd.DataFrame([current_prices])
        
        if os.path.exists(file_path):
            # 기존 파일이 있으면 데이터를 읽어옴
            try:
                existing_df = pd.read_excel(file_path)
                # 기존 데이터와 새로운 데이터를 합침
                df = pd.concat([existing_df, new_df], ignore_index=True)
            except Exception as e:
                print(f"기존 파일 읽기 실패: {e}")
                df = new_df
        else:
            df = new_df

        # 날짜로 정렬
        df['날짜 및 시간'] = pd.to_datetime(df['날짜 및 시간'])
        df = df.sort_values('날짜 및 시간')
        df['날짜 및 시간'] = df['날짜 및 시간'].dt.strftime('%Y-%m-%d %H:%M')

        # 데이터 저장
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='w') as writer:
            df.to_excel(writer, index=False, sheet_name=pcode)

        print("결과가 'danawa_lowest_prices_" + pcode + ".xlsx' 파일로 저장되었습니다.")

        # 차트 생성
        wb = load_workbook(file_path)
        ws = wb.active

        # 기존 차트 제거
        for chart in ws._charts:
            ws.remove_chart(chart)

        # 새 차트 생성
        chart = LineChart()
        chart.title = "Mall별 가격 변화"
        chart.style = 13
        chart.y_axis.title = '가격'
        chart.x_axis.title = '날짜 및 시간'
        chart.width = 40  
        chart.height = 20  

        # 데이터 참조
        data = Reference(ws, min_col=2, min_row=1, max_col=ws.max_column, max_row=ws.max_row)
        categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        # 차트 설정
        colors = ["FF0000", "00FF00", "0000FF", "FFFF00", "FF00FF", "00FFFF", "000000", "800000", "008000", "000080"]
        for i, series in enumerate(chart.series):
            series.graphicalProperties.line.width = 25000
            series.graphicalProperties.line.solidFill = colors[i % len(colors)]
            series.smooth = True
            series.marker.symbol = "circle"
            series.marker.size = 5
            series.marker.graphicalProperties.solidFill = colors[i % len(colors)]
            series.marker.graphicalProperties.line.solidFill = colors[i % len(colors)]

        # 차트 추가
        ws.add_chart(chart, "O5")  # 차트를 H5 위치에 추가하여 오른쪽으로 이동

        # 엑셀 파일 저장
        wb.save(file_path)
        print("차트가 'danawa_lowest_prices_" + pcode + ".xlsx' 파일에 추가되었습니다.")

crawl_and_update_excel()